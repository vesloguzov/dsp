import os
from openpyxl import Workbook, load_workbook
from datetime import date, datetime
import json

def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""

    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError("Type %s not serializable" % type(obj))


main_wb = load_workbook('/home/vesloguzov/PROJECTS/dsp/dsp/Курсы_в_РОО_экспертиза_описания.xlsx', data_only=True)

main_ws = main_wb[main_wb.sheetnames[2]]
main_ws = main_wb.get_sheet_by_name('Чеклист')

print(main_ws.max_row)
ex_list = []


def get_bool_value(cell_value):
    # print(cell_value)
    if str(cell_value).strip() == "1.0" or str(cell_value).strip() == "1":
        return True
    else:
        return False

gipermet_count = 0
for ex_num in range(2, 727):
    # print(ex_num)

    if main_ws[f"D{ex_num}"].value.strip() != "СЭО УрФУ на платформе Гиперметод":
        ex_list.append({
            "course_title": main_ws[f"H{ex_num}"].value,
            "course_institution": main_ws[f"G{ex_num}"].value,
            "course_partner": main_ws[f"D{ex_num}"].value,
            "external_url": main_ws[f"E{ex_num}"].value,
            # "ex_date": main_ws[f"C{ex_num}"].value,
            # "type": "0",
            # "executed": True,
            # "expert": main_ws[f"B{ex_num}"].value,
            # "supervisor": "",
            # "organizer": "ИТОО",
            # "comment": main_ws[f"AP{ex_num}"].value,
            # "comment_fieldset_1": "",
            # "comment_fieldset_2": "",
            #
            # "has_length": get_bool_value(main_ws[f"I{ex_num}"].value),
            # "has_description": get_bool_value(main_ws[f"J{ex_num}"].value),
            # "has_authors": get_bool_value(main_ws[f"K{ex_num}"].value),
            # "language": main_ws[f"L{ex_num}"].value,
            # "has_prerequisites": get_bool_value(main_ws[f"M{ex_num}"].value),
            # "has_certificate": get_bool_value(main_ws[f"N{ex_num}"].value),
            # "has_dates": get_bool_value(main_ws[f"O{ex_num}"].value),
            # "has_admin_email": get_bool_value(main_ws[f"P{ex_num}"].value),
            # "has_labor": get_bool_value(main_ws[f"Q{ex_num}"].value),
            # "has_competences": get_bool_value(main_ws[f"R{ex_num}"].value),
            # "has_results": get_bool_value(main_ws[f"S{ex_num}"].value),
            # "has_evaluation_tools": get_bool_value(main_ws[f"T{ex_num}"].value),
            # "has_recommended_directions": get_bool_value(main_ws[f"U{ex_num}"].value),
            # "has_proctoring": get_bool_value(main_ws[f"V{ex_num}"].value),
            # "has_labor_costs": get_bool_value(main_ws[f"W{ex_num}"].value),
            # "has_short_description": get_bool_value(main_ws[f"X{ex_num}"].value),
            # "has_learning_plan": get_bool_value(main_ws[f"Y{ex_num}"].value),
            # "has_promo_clip": get_bool_value(main_ws[f"Z{ex_num}"].value),
            # "language_video": main_ws[f"AA{ex_num}"].value,
            # "language_subtitles": main_ws[f"AB{ex_num}"].value,
            # "has_course_subject": get_bool_value(main_ws[f"AC{ex_num}"].value),
            # "is_open": get_bool_value(main_ws[f"AD{ex_num}"].value),
            # "has_expertises_types": get_bool_value(main_ws[f"AE{ex_num}"].value),
            # "has_ownership_document_scan": get_bool_value(main_ws[f"AF{ex_num}"].value),
            # "has_not_prohibited": get_bool_value(main_ws[f"AG{ex_num}"].value),
            # "has_text_materials": get_bool_value(main_ws[f"AH{ex_num}"].value),
            # "has_illustrations": get_bool_value(main_ws[f"AI{ex_num}"].value),
            # "has_audio": get_bool_value(main_ws[f"AJ{ex_num}"].value),
            # "has_video": get_bool_value(main_ws[f"AK{ex_num}"].value),
            # "has_quality_checking": main_ws[f"AL{ex_num}"].value,
            # "no_permission_of_owners": main_ws[f"AM{ex_num}"].value,
            # "got_into_record": main_ws[f"AN{ex_num}"].value,
            # "got_expertise_2018": main_ws[f"AO{ex_num}"].value,
            # "additional_info": main_ws[f"AQ{ex_num}"].value
        })
    else:
        print(main_ws[f"H{ex_num}"].value)
        gipermet_count += 1

print("gipermet_count: ", gipermet_count)

with open("dump.json", "w") as dump:
    dump.write(json.dumps(ex_list, default=json_serial))

print("complete")
#         # Passport
#           has_length = models.BooleanField("Длительность", default=False)
#           has_description = models.BooleanField("Описание", default=False)
#           has_authors = models.BooleanField("Авторы", default=False)
#           language = models.CharField("Язык содержания", default="русский", max_length=255)
#           has_prerequisites = models.BooleanField("Рекомендуемые \"входные\" требования к обучающемуся", default=False)
#           has_certificate = models.BooleanField("Сертификат (выдается или нет)", default=False)
#           has_dates = models.BooleanField("Даты ближайшего запуска", default=False)
#           has_admin_email = models.BooleanField("Адрес эл. почты администратора ОК", default=False)
#           has_labor = models.BooleanField("Трудоемкость", default=False)
#           has_competences = models.BooleanField("Компетенции", default=False)
#           has_results = models.BooleanField("Результаты обучения", default=False)
#           has_evaluation_tools = models.BooleanField("Оценочные средства", default=False)
#           has_recommended_directions = models.BooleanField("Рекомендуемые направления подготовки", default=False)
#           has_proctoring = models.BooleanField("Наличие подтвержденного сертификата (сервис прокторинга)", default=False)
#           has_labor_costs = models.BooleanField("Трудозатраты", default=False)
#           has_short_description = models.BooleanField("Короткое описание", default=False)
#           has_learning_plan = models.BooleanField("Учебный план", default=False)
#           has_promo_clip = models.BooleanField("Проморолик", default=False)
#           language_video = models.CharField("Язык видео", default="русский", max_length=255)
#           language_subtitles = models.CharField("Язык субтитров", default="русский", max_length=255)
#           has_course_subject = models.BooleanField("Предмет курса", default=False)
#           is_open = models.BooleanField("Открытость курса", default=False)
#           has_expertises_types = models.BooleanField("Типы экспертиз для допуска", default=False)
#           has_ownership_document_scan = models.BooleanField("Скан документа, подтверждающего правообладание")
#           has_not_prohibited = models.BooleanField("В курсе отсутствуют запрещенные материалы", default=False)
#           has_text_materials = models.BooleanField("Текстовые материалы", default=False)
#           has_illustrations = models.BooleanField("Иллюстрации", default=False)
#           has_audio = models.BooleanField("Аудиоматериалы", default=False)
#           has_video = models.BooleanField("Видеоматериалы", default=False)

#         has_quality_checking = models.BooleanField("прошел проверку обязательной оценки качества", default=False)

#           got_into_record = models.CharField("попал в отчет", max_length=255, null=True, blank=True)
#           got_expertise_2018 = models.BooleanField("прошел экспертизу в 2018 (1 квартал)", default=False)
#           additional_info = models.TextField("Дополнительная информация", null=True, blank=True)

#
# def range_values_array(ws, range):
#     arr = []
#     rows = ws[range]
#     for row in rows:
#         for cell in row:
#             arr.append(cell.value)
#     return arr
#
# ExpertiseList = []
#
# is_end = False
# possible_end = 0
# cell_num = 2
# vals = []
# while not is_end:
#     #
#     Expertise = []
#     for a in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ"]:
#         Expertise.append(main_ws[f"{a}{cell_num}"].value)
#     #
#     if Expertise[1]:
#         vals.append({

#     })
#         "title": course[1],
#         "expertise_status": course[2],
#         "activities": course[3],
#         "roo_status": course[5],
#         "platform": course[6],
#         "course_url": course[7],
#         "owner": course[8],
#         "state": course[9],
#         "date": course[10],
#         "communication_owner": course[11],
#         "expertise_types": course[12],
#         "expertise_passed": course[13],
#         "expert": course[14],
#         "expert_login": course[15],
#         "supervisor": course[16],
#         "organizer": course[17],
#         "contacts": "/n".join(["" if not n else n for n in course[18:20]])
#     #
#     cell_num += 1
#     #
#     if Expertise[0] is not None:
#         cell_num += 1
#         possible_end = 0
#     else:
#         possible_end += 1
#         cell_num += 1
#         if possible_end == 30:
#             is_end = True
# #
# import json
#
# with open("dump.json", "w") as dump:
#     dump.write(json.dumps(vals, default=json_serial))
